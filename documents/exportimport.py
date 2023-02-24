# Python
from datetime import datetime
import pytz
# Django Classes
from pathlib import Path
from django.contrib.auth.models import Group
# Openpyxl Classes
from openpyxl import load_workbook
# My Classes
from accounts.models import CustomUser
from locations.models import *
from dockingRecords.models import *
from django.contrib.auth.hashers import make_password


# Build paths inside the project like this: BASE_DIR / 'subdir'.
DOCS_DIR = Path(__file__).resolve().parent.parent
    
# Set timezone to Riyadh
tz = pytz.timezone('Asia/Riyadh')

################################################################################################################################################################################
############################################################################# Accounts #########################################################################################
################################################################################################################################################################################
    
# take info from the excel sheet and create user info from it
def create_account_fields(sheet_obj ,i):
    user_id    = str(sheet_obj.cell(row = i, column = 2).value)
    first_name = str(sheet_obj.cell(row = i, column = 3).value)
    last_name  = str(sheet_obj.cell(row = i, column = 4).value)
    username   = str(sheet_obj.cell(row = i, column = 5).value)
    password = str(sheet_obj.cell(row = i, column = 6).value)
    email      = str(sheet_obj.cell(row = i, column = 7).value)
    terminal   = Terminal.objects.get(id = sheet_obj.cell(row = i, column = 8).value)
    active     = True if sheet_obj.cell(row = i, column = 10).value == 'True' else False

    return user_id, first_name, last_name, username, password, email, terminal, active

# Add groups to the created users
#def add_passwords(max_row, sheet_obj):
#    for i in range(2, max_row + 1):
#        password = str(sheet_obj.cell(row = i, column = 6).value)
#        user_id  = sheet_obj.cell(row = i, column = 2).value
#        user     = CustomUser.objects.get(user_id = user_id)
#        user.set_password(password)

# Add groups to the created users
def add_groups(max_row, sheet_obj):
    for i in range(2, max_row + 1):
        user_id = sheet_obj.cell(row = i, column = 2).value
        groups  = Group.objects.get(id = sheet_obj.cell(row = i, column = 9).value)
        user    = CustomUser.objects.get(user_id = user_id)
        user.groups.add(groups)

# copy users to SeconOperator Model in dockingRecords
def copy_accounts(max_row, sheet_obj):
    copy_acc_list = []

    for i in range(2, max_row + 1):
        user_id = sheet_obj.cell(row = i, column = 2).value
        user    = CustomUser.objects.get(user_id = user_id)
        
        copy_acc_list.append(SecondOperator(second_operator = user))
    
    SecondOperator.objects.bulk_create(copy_acc_list)

# bulk create create accounts from the excel sheet provided
def create_accounts(file_name):
    path      = f'{DOCS_DIR}/media/{file_name}'
    wb_obj    = load_workbook(path)
    sheet_obj = wb_obj.active
    max_row   = sheet_obj.max_row

    # list to save user instances in then bulk create them saves too much time
    acc_list = []

    # loop through the sheet
    for i in range(2, max_row + 1):
        user_id, first_name, last_name, username, password, email, terminal, active = create_account_fields(sheet_obj ,i)
        # continue if user exists
        if CustomUser.objects.filter(email = email) or CustomUser.objects.filter(username = username) or CustomUser.objects.filter(user_id = user_id):
            continue
        acc_list.append(CustomUser(user_id = user_id,
                first_name = first_name,
                last_name = last_name,
                username = username,
                password = make_password(password),
                email = email,
                location = terminal,
                is_active = active
            )
        )

    # bulk_create the accounts from the list
    created_accounts = CustomUser.objects.bulk_create(acc_list)

    # add the groups manually becuase it's a many to many relationship and that doesn't work with bulk create
    add_groups(max_row, sheet_obj)
    copy_accounts(max_row, sheet_obj)
    
    return len(created_accounts)

################################################################################################################################################################################
############################################################################# docking Records ##################################################################################
################################################################################################################################################################################

# format datetime fields in django datetime format
def format_datetimefields(datetime_field):
    try:
        # get docking details
        datetime_field = datetime_field.replace('-','/')
        # convert it to this format
        datetime_field = datetime.strptime(datetime_field, "%Y/%m/%d %H:%M:%S")
        # Include timezone info in it
        datetime_field = datetime.astimezone(datetime_field,tz=tz)
    except:
        return None

    return datetime_field

# format time fields in django time format
def format_timefields(time_field):
    try:
        time_field = time_field.replace('-','/')
        time_field = datetime.strptime(time_field, "%H:%M:%S")
        time_field = datetime.astimezone(time_field, tz=tz)
    except:
        return None

    return time_field

# take info from the excel sheet and create record info from it
def create_record_fields(sheet_obj ,i):
    no_docking = True if sheet_obj.cell(row = i, column = 1).value == "True" else False
    operator1  = CustomUser.objects.get(username = sheet_obj.cell(row = i, column = 2).value)
    stand      = str(sheet_obj.cell(row = i, column = 3).value)
    if '14' in stand:
        stand = '15B'
    if '22' in stand:
        stand = '23B'
    if '24' in stand:
        stand = '25B'
    flight_no1 = sheet_obj.cell(row = i, column = 4).value
    ac_type    = sheet_obj.cell(row = i, column = 5).value
    ac_reg     = sheet_obj.cell(row = i, column = 6).value
    docked     = format_datetimefields(f'{sheet_obj.cell(row = i, column = 8).value} {sheet_obj.cell(row = i, column = 9).value}')
    chocks_on  = format_timefields(sheet_obj.cell(row = i, column = 7).value)
    b_used     = False
    b_docked   = None
    operator2  = SecondOperator.objects.get(second_operator = operator1)
    flight_no2 = sheet_obj.cell(row = i, column = 12).value
    door_close = format_timefields(sheet_obj.cell(row = i, column = 13).value)
    undocked   = format_datetimefields(f'{sheet_obj.cell(row = i, column = 8).value} {sheet_obj.cell(row = i, column = 14).value}')
    b_undocked = None
    remarks    = f'Arrival Remarks: {sheet_obj.cell(row = i, column = 10).value} - {sheet_obj.cell(row = i, column = 11).value} - Departure Remarks: {sheet_obj.cell(row = i, column = 15).value} - {sheet_obj.cell(row = i, column = 16).value}'

    return no_docking, operator1, stand, flight_no1, ac_type, ac_reg, docked, chocks_on, remarks, b_used, b_docked, operator2, flight_no2, door_close, undocked, b_undocked

def record_exists(operator1, stand, flight_no1, ac_type, docked):
    # doesn't work if dublicate data already exists
    # check if a record exists
    try:
        record_exists = DockingRecord.objects.get(operator1 = operator1, stand = stand, flight_no1 = flight_no1,ac_type = ac_type, docked = docked)
        return True
    except:
        return False


def create_records_data(file_name):
    path      = f'{DOCS_DIR}/media/{file_name}'
    wb_obj    = load_workbook(path)
    sheet_obj = wb_obj.active
    max_row   = sheet_obj.max_row

    skip_iter = False
    record_list = []

    for i in range(2, max_row + 1):
        # initiate variables
        no_docking, operator1, stand,\
        flight_no1, ac_type, ac_reg,\
        docked, chocks_on, remarks,\
        b_used, b_docked, operator2,\
        flight_no2, door_close,\
        undocked, b_undocked = create_record_fields(sheet_obj ,i)

        # skip this iteration if b_used was true in the previous iteration
        if skip_iter:
            # reset variable
            skip_iter = False
            continue

        # in case bravo used we dont need two records anymore so we are unifying the data
        if "B" in stand:
            # pbb bravo used
            b_used = True
            # get bravo docking and undocking time
            b_docked   = format_timefields(sheet_obj.cell(row = i, column = 9).value)
            b_undocked = format_timefields(sheet_obj.cell(row = i, column = 14).value)
            # get Alpha docking and undocking time from the next row
            docked   = format_datetimefields(f'{sheet_obj.cell(row = i+1, column = 8).value} {sheet_obj.cell(row = i+1, column = 9).value}')
            undocked = format_datetimefields(f'{sheet_obj.cell(row = i+1, column = 8).value} {sheet_obj.cell(row = i+1, column = 14).value}')
            # skip next row after the start of the next iteration starts if the next row has the same flight number
            if flight_no1 == sheet_obj.cell(row = i+1, column = 4).value:
                skip_iter = True
            else:
                docked   = format_datetimefields(f'{sheet_obj.cell(row = i, column = 8).value} {sheet_obj.cell(row = i+1, column = 9).value}')
                undocked = format_datetimefields(f'{sheet_obj.cell(row = i, column = 8).value} {sheet_obj.cell(row = i+1, column = 14).value}')


        # remove A, or B from stand name then get stand
        stand = Stand.objects.get(stand_name = stand[:2])

        # returns true if a record exists 
        if record_exists(operator1, stand, flight_no1, ac_type, docked):
            continue

        # if True don't save undocking data
        if no_docking:
            record_list.append(DockingRecord(
                no_docking = no_docking,
                operator1 = operator1,
                stand = stand,
                flight_no1 = flight_no1,
                ac_type = ac_type,
                ac_reg = ac_reg,
                docked = docked,
                chocks_on = chocks_on,
                remarks = remarks,
                b_used = b_used,
                b_docked = b_docked
                )
            )
            continue

        # create a new record an append it to the list
        record_list.append(DockingRecord(
            no_docking = no_docking,
            operator1 = operator1,
            stand = stand,
            flight_no1 = flight_no1,
            ac_type = ac_type,
            ac_reg = ac_reg,
            docked = docked,
            chocks_on = chocks_on,
            remarks = remarks,
            b_used = b_used,
            b_docked = b_docked,
            operator2 = operator2,
            flight_no2 = flight_no2,
            door_close = door_close,
            undocked = undocked,
            b_undocked = b_undocked
            )
        )

    # bulk create all new records using the list
    objects_created = DockingRecord.objects.bulk_create(record_list)
    return len(objects_created)



# could be useful for something else

    #def create_new_record(no_docking, operator1, stand, flight_no1, ac_type, ac_reg, chocks_on, docked, remarks, b_used, b_docked, operator2, flight_no2, door_close, undocked, b_undocked):
#    # check if a record exists
#    try:
#        record_exists = DockingRecord.objects.get(operator1 = operator1, stand = stand, flight_no1 = flight_no1,ac_type = ac_type, docked = docked)
#        return False
#    # create a record
#    except:
#        new_record, record_created = DockingRecord.objects.get_or_create(
#            no_docking = no_docking,
#            operator1 = operator1,
#            stand = stand,
#            flight_no1 = flight_no1,
#            ac_type = ac_type,
#            ac_reg = ac_reg,
#            docked = docked,
#            chocks_on = chocks_on,
#            remarks = remarks,
#            b_used = b_used,
#            b_docked = b_docked,
#            operator2 = operator2,
#            flight_no2 = flight_no2,
#            door_close = door_close,
#            undocked = undocked,
#            b_undocked = b_undocked,
#    )
## unused fields in Dammam project
#                c_used,
#                c_docked,
#                c_undocked,
#    return record_created

################################################################################################################################################################################
############################################################################# locations ########################################################################################
################################################################################################################################################################################

#def handle_uploaded_file(f):  
#    with open('ProjectManager/static/upload/'+f.name, 'wb+') as destination:  
#        for chunk in f.chunks():  
#            destination.write(chunk)

# slow but insures no doublicate data can be created because that creates conflicts in the database
def create_gates(max_row, sheet_obj):
    gates_list = []
    for i in range(2, max_row + 1):
        terminal = Terminal.objects.get(id = sheet_obj.cell(row = i, column = 1).value)
        gate     = sheet_obj.cell(row = i, column = 2).value

        # Prevents double imports for the same data
        try:
            gate = Gate.objects.get(terminal_name = terminal, gate_name = gate)
            continue
        except:
            gate, gate_created = Gate.objects.get_or_create(terminal_name = terminal, gate_name = gate)
            gates_list.append(gate)

    created_gates = len(gates_list)
    return created_gates

# slow but insures no doublicate data can be created because that creates conflicts
def create_stands(max_row, sheet_obj):
    stands_list = []
    for i in range(2, max_row + 1):
        gate  = Gate.objects.get(gate_name = sheet_obj.cell(row = i, column = 2).value)
        stand = sheet_obj.cell(row = i, column = 3).value

        try:
            stand = Stand.objects.get(gate_name = gate, stand_name = stand)
            continue
        except:
            stand, stand_created = Stand.objects.get_or_create(gate_name = gate, stand_name = stand)
            stands_list.append(stand)
    
    created_stands = len(stands_list)
    return created_stands


def create_pbbs(max_row, sheet_obj):
    pbbs_list = []
    for i in range(2, max_row + 1):
        stand = Stand.objects.get(stand_name = sheet_obj.cell(row = i, column = 3).value)
        pbb1  = sheet_obj.cell(row = i, column = 4).value
        # returns None if cell is empty
        pbb2  = sheet_obj.cell(row = i, column = 5).value
        pbb3  = sheet_obj.cell(row = i, column = 6).value

        try:
            pbb = Pbb.objects.get(stand_name = stand, pbb_name = pbb1)
        except:
            pbbs_list.append(Pbb(stand_name = stand, pbb_name = pbb1))

        if pbb2 != None:
            try:
                pbb = Pbb.objects.get(stand_name = stand, pbb_name = pbb2)
            except:
                pbbs_list.append(Pbb(stand_name = stand, pbb_name = pbb2))

        if pbb3 != None:
            try:
                pbb = Pbb.objects.get(stand_name = stand, pbb_name = pbb3)
            except:
                pbbs_list.append(Pbb(stand_name = stand, pbb_name = pbb3))
    
    created_pbbs = len(Pbb.objects.bulk_create(pbbs_list))
    return created_pbbs


def create_location_data(file_name):
    path      = f'{DOCS_DIR}/media/{file_name}'
    wb_obj    = load_workbook(path)
    sheet_obj = wb_obj.active
    max_row   = sheet_obj.max_row

    created_gates  = create_gates(max_row, sheet_obj)
    created_stands = create_stands(max_row, sheet_obj)
    created_pbbs   = create_pbbs(max_row, sheet_obj)

    return created_gates, created_stands, created_pbbs