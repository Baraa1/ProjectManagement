# Django
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse
from pathlib import Path
from django.contrib.auth.decorators import login_required, user_passes_test
# Openpyxl
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
# Custom
from .models import Document
from .forms import DocumentForm
from .exportimport import *
from ProjectManager.decorators import *
from accounts.decorators import *
from locations.decorators import *
from dockingRecords.decorators import *
# Create your views here.

################################################################################################################################################################################
############################################################################# Accounts #########################################################################################
################################################################################################################################################################################

@login_required(login_url='login')
@user_passes_test(admin)
@user_passes_test(add_accounts)
def upload_accounts_file(request):
    # Handle file upload
    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.accounts_file = True
            instance.save()
            # Redirect to the document list after POST
            return redirect('upload_accounts_file')
    else:
        form = DocumentForm() # A empty, unbound form

    # Load documents for the list page
    documents = Document.objects.all()
    
    context = {
        "documents":documents,
        "form":form,
        "accounts_view": True,
    }

    # Render list page with the documents and the form
    return render(request, 'accountsexportimport.html',context)

@login_required(login_url='login')
@user_passes_test(admin)
@user_passes_test(add_accounts)
def delete_accounts_file(request,file_id):
    file = get_object_or_404(Document, id = file_id)
    file.delete()
    messages.success(request, f'{file.docfile.name} Deleted')
    return redirect('upload_accounts_file')

@login_required(login_url='login')
@user_passes_test(admin)
@user_passes_test(add_accounts)
def import_accounts_data(request, file_id):
    file_name = Document.objects.get(id = file_id).docfile.name
    number_of_accounts = create_accounts(file_name)
    messages.success(request, f'{number_of_accounts} Accounts were created')
    return redirect('upload_accounts_file')

@login_required(login_url='login')
@user_passes_test(admin)
@user_passes_test(add_accounts)
def acc_template(request):
    # Get the file
    DOCS_DIR = Path(__file__).resolve().parent.parent.parent
    path = f'{DOCS_DIR}/Docs/Accounts.xlsx'
    wb_obj   = load_workbook(path)
    # returns a file to the user 
    response = HttpResponse(content=save_virtual_workbook(wb_obj), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Accounts.xlsx'
    
    return response

################################################################################################################################################################################
############################################################################# docking Records ##################################################################################
################################################################################################################################################################################

@login_required(login_url='login')
@user_passes_test(admin)
@user_passes_test(add_records)
def upload_records_file(request):
    # Handle file upload
    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.docking_records_file = True
            instance.save()
            # Redirect to the document list after POST
            return redirect('upload_records_file')
    else:
        form = DocumentForm() # A empty, unbound form

    # Load documents for the list page
    documents = Document.objects.all()
    
    context = {
        "documents":documents,
        "form":form,
        "docking_records_view": True,
    }

    # Render list page with the documents and the form
    return render(request, 'recordsexportimport.html',context)

@login_required(login_url='login')
@user_passes_test(admin)
@user_passes_test(add_records)
def delete_records_file(request,file_id):
    file = get_object_or_404(Document, id = file_id)
    file.delete()
    messages.success(request, f'{file.docfile.name} Deleted')
    return redirect('upload_records_file')

@login_required(login_url='login')
@user_passes_test(add_records)
def import_records_data(request, file_id):
    file_name       = Document.objects.get(id = file_id).docfile.name
    objects_created = create_records_data(file_name)
    messages.info(request, f'{objects_created} records were successfully Imported')
    return redirect('upload_records_file')

@login_required(login_url='login')
@user_passes_test(admin)
@user_passes_test(add_records)
def rec_template(request):
    # Get the file
    DOCS_DIR = Path(__file__).resolve().parent.parent.parent
    path = f'{DOCS_DIR}/Docs/Docking-records.xlsx'
    wb_obj   = load_workbook(path)
    # creates a file response for the user 
    response = HttpResponse(content=save_virtual_workbook(wb_obj), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Docking-records.xlsx'
    
    return response

################################################################################################################################################################################
############################################################################# Locations ########################################################################################
################################################################################################################################################################################

@login_required(login_url='login')
@user_passes_test(admin)
def upload_location_file(request):
    # Handle file upload
    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.locations_file = True
            instance.save()
            # Redirect to the document list after POST
            return redirect('upload_location_file')
    else:
        form = DocumentForm() # A empty, unbound form

    # Load documents for the list page
    documents = Document.objects.all()
    
    context = {
        "documents":documents,
        "form":form,
        "locations_view": True,
    }

    # Render list page with the documents and the form
    return render(request, 'exportimport.html',context)

@login_required(login_url='login')
@user_passes_test(admin)
def delete_location_file(request,file_id):
    file = get_object_or_404(Document, id = file_id)
    file.delete()
    messages.success(request, f'{file.docfile.name} Deleted')
    return redirect('upload_location_file')

@login_required(login_url='login')
@user_passes_test(admin)
def import_location_data(request, file_id):
    file_name = Document.objects.get(id = file_id).docfile.name
    created_gates, created_stands, created_pbbs = create_location_data(file_name)
    messages.info(request, f'{created_gates} Gates, {created_stands} Stands and {created_pbbs} Pbbs were created')
    return redirect('upload_location_file')

@login_required(login_url='login')
@user_passes_test(admin)
def locations_template(request):
    # Get the file
    DOCS_DIR = Path(__file__).resolve().parent.parent.parent
    path = f'{DOCS_DIR}/Docs/Location.xlsx'
    wb_obj   = load_workbook(path)
    # returns a file to the user 
    response = HttpResponse(content=save_virtual_workbook(wb_obj), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Location-template.xlsx'
    
    return response
